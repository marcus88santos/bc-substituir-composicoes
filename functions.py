from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv
load_dotenv()
import io

def read_sheet(file):
  worksheet = load_workbook(file)
  sheet = worksheet.active
  
  changes = []
  cols = {'status': '', 'codigo_a_substituir': '', 'base': '', 'novo_tipo': '', 'novo_codigo': '', 'nova_base': ''}
  
  # iterando sobre o cabeçalho
  for row in sheet.iter_rows(1,1):
    for cel in row:
      for col in cols:
        if cel.value == col:
          cols[col] = get_column_letter(cel.column)
  
  # iterando sobre os dados
  for cel in sheet[f'${cols['status']}']:
    if cel.value == 'ativo':
      # iterando sobre uma substituição com status 'ativo'
      for row in sheet.iter_rows(cel.row,cel.row):
        dict = {}
        for cel in row:
          for col in cols:
            # verificando se a coluna é a mesma do cabeçalho
            if col != 'status' and get_column_letter(cel.column) == cols[col]:
              dict[col] = cel.value
        changes.append(dict)
  
  return changes

def get_changes(folder, file):
  old_os_folder = os.getcwd()
  user = os.getcwd().replace('\\', '/')
  user = user[0:user.find('/', user.find('/', user.find('/') + 1) + 1)]
  new_os_folder = user + folder
  os.chdir(new_os_folder)
  
  changes = read_sheet(file)
  os.chdir(old_os_folder)
  
  return changes

def get_orcamento(folder, title):
  old_os_folder = os.getcwd()
  user = os.getcwd().replace('\\', '/')
  user = user[0:user.find('/', user.find('/', user.find('/') + 1) + 1)]
  new_os_folder = user + folder
  os.chdir(new_os_folder)
  
  orcamento = io.open(title, mode="r", encoding='utf-8').read()
  
  os.chdir(old_os_folder)
  
  return orcamento